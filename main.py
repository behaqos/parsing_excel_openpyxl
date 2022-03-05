from openpyxl import load_workbook

#TODO добавить кабинеты
#TODO добавить дату с днём недели (1-ая колонка) и время из 2-ой колонки, если есть



'''
    Проходимся по всем листам книги
    А в каждом листе, начиная именно с третьего (т.к. 1 и 2 это дата и время) проходимся по названиям групп
    Если название группы совпадает с целевой, то
'''
def search_sheet(class_group, book):
    for sheet in book:
        for column_number in range(3, 26):
            string_value = sheet.cell(row=16, column=column_number).value
            if class_group == string_value:
                # if string_value != None and string_value != 'КАБ':
                print("Результат:            " + string_value)
                print("Колонка группы в Excel:      №" + column_number.__str__())
                cabinet_column_number = column_number+1
                print("Колонка кабинета в Excel:    №" + (column_number+1).__str__())


                result_sheet = sheet
                return result_sheet, column_number, cabinet_column_number

#FIXME остановился над вопросом как выести значения колонки группы, а затем нужно вывести с ним же время, дату и кабинет, которые находятся в колонках по соседству на одной строке

'''
Сначала берёт колонку и переводит из цифры в букву. Далее берёт границы,
которые во всех листах начинаются от 21 до 50.
А дальше проходит по листу.
'''


'''
Здесь row, который внутри типа данных кортеж содержит все данные столбца.

Можно все необходимые столбцы скопировать в новую структуру данных, а оттуда
потом и печатать.
Разуменее здесь сначала брать одно поле, а затем по полю отбирать день + предмет + кабинет.
Потому что соединив по отдельности дату, время, предмет и кабинет, соотносить их будет нереально.
'''
def search_schedule(sheet, column_number, cabinet_column_number):
    subject_column = chr(64 + column_number) # переводим колонку в буквенное значение
    rows_range =  subject_column + '21' + ':' + subject_column + '50' # берём отрезок данных
    # print(rows_range)

    date_column = 'A'
    date_rows_range = date_column + '21' + ':' + date_column + '50'

    cabinet_column = chr(64 + cabinet_column_number)
    cabinet_rows_range = cabinet_column + '21' + ':' + subject_column + '50'
    subjects = sheet[rows_range]
    dates = sheet[date_rows_range]
    cabinets = sheet[cabinet_rows_range]
    # print(sheet[cabinet_rows_range][1].value)

    for row in sheet[date_rows_range]:
        for r in row:
            if r.value != None:
                print(r.value)

if __name__ == '__main__':
    book = load_workbook(filename='/Users/dpenney/Downloads/расписание.xlsx')
    # class_group = input("Какой у тебя класс?\n")
    # Reading all sheets

    class_group = "ИСП/Р-21-11"
    result_sheet = None
    print("Начинаю поиск:        " + class_group)

    result_sheet, column_number, cabinet_column_number = search_sheet(class_group, book)
    print("\nГруппа " + class_group + " находится в книге ", end=' ')
    print(result_sheet)
    print("\nДальше ищём занятия группы " + class_group + '\n')

    search_schedule(result_sheet, column_number, cabinet_column_number)


# See PyCharm help at https://www.jetbrains.com/help/pycharm/
